from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, Http404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET, require_POST
from django_ratelimit.decorators import ratelimit
from functools import wraps

from .forms import AlexSearchForm, AlexOrderForm
from .models import AlexSearchJob, AlexOrder
from .services.job_runner import run_scraping_job


def _safe_filename(job, ext: str) -> str:
    """openalex_<arama_kelimesi>_<YYYYMMDD>.<ext> formatında güvenli dosya adı."""
    import re
    from django.utils import timezone
    keyword = ''
    if job.query_parts:
        for part in job.query_parts:
            val = part.get('value', '').strip()
            if val:
                keyword = val
                break
    keyword = (keyword or 'sonuclar')[:40]
    keyword = re.sub(r'[^\w\s-]', '', keyword, flags=re.UNICODE)
    keyword = re.sub(r'[\s]+', '_', keyword).strip('_')
    date_str = timezone.now().strftime('%Y%m%d')
    return f'openalex_{keyword}_{date_str}.{ext}'


def feature_required(flag_name):
    """Decorator: SiteSettings'deki feature flag kapalıysa 404 döner."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            from forum.models import SiteSettings
            site = SiteSettings.load()
            if not getattr(site, f'feature_{flag_name}', True):
                raise Http404
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


IBAN_INFO = {
    'hesap_sahibi': 'Bünyamin Keleş',
    'iban': 'TR73 0003 2000 0000 0079 1034 65',
}


@feature_required('openalex')
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def openalex_landing(request):
    """Landing page: gelişmiş arama formu + demo arama."""
    if not request.user.is_authenticated:
        return render(request, 'service_promo.html', {
            'promo_title': 'OpenAlex Yayın Kazıma ve Veri İndirme Aracı',
            'promo_icon': 'bi-search',
            'promo_color': 'primary',
            'promo_description': '240 milyondan fazla akademik yayından kodsuz veri kazıma aracı. Başlık, yazar, kurum gibi alanlarda arama yapın; sonuçları Excel veya TXT olarak tek tıkla indirin.',
            'promo_features': [
                {'icon': 'bi-download', 'title': 'Excel & TXT İndirme', 'desc': 'Yayın başlığı, yazar, dergi, yıl, DOI ve özet verilerini tek tıkla Excel veya TXT olarak indirin.'},
                {'icon': 'bi-database-fill', 'title': '240M+ Kaynaktan Veri Çekme', 'desc': 'OpenAlex\'in tüm akademik veri tabanından makale, kitap ve konferans bildirisini kodsuz kazıyın.'},
                {'icon': 'bi-sliders', 'title': 'Kodsuz Gelişmiş Sorgulama', 'desc': 'Başlık, özet, yazar, dergi, kurum, yıl ve DOI gibi 9 farklı alanda AND/OR sorguları oluşturun.'},
                {'icon': 'bi-bar-chart-line-fill', 'title': 'Tek Tıkla Bibliometrik Analiz', 'desc': '100+ sonuçta tek tıkla bibliometrik analize gönderin, PDF rapor alın.'},
            ],
            'promo_steps': [
                'Arama kriterlerinizi seçin: başlık, yazar, kurum, yıl veya DOI.',
                'Sistem OpenAlex\'ten verileri kazıyarak saniyeler içinde listeler.',
                'Excel veya TXT olarak indirin ya da bibliometrik analiz başlatın.',
            ],
        })
    form = AlexSearchForm()
    user = request.user

    daily_count = AlexSearchJob.daily_count_for_user(user)
    daily_limit = AlexSearchJob.get_daily_limit(user)
    remaining = max(0, daily_limit - daily_count)

    if request.method == 'POST':
        form = AlexSearchForm(request.POST)
        if form.is_valid():
            if not user.profile.email_verified:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'error': 'OpenAlex tarama için e-posta doğrulaması gereklidir. Profil sayfanızdan e-postanızı doğrulayın.'
                    }, status=403)
                return render(request, 'openalex/landing.html', {
                    'form': form,
                    'error': 'OpenAlex tarama için e-posta doğrulaması gereklidir.',
                    'remaining': remaining,
                    'daily_limit': daily_limit,
                })

            if remaining <= 0:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'error': f'Günlük demo limitiniz doldu ({daily_limit}/{daily_limit}). '
                                 f'{"Premium üyelikle 7 aramaya yükseltebilirsiniz." if not user.profile.is_premium else "Yarın tekrar deneyebilirsiniz."}'
                    }, status=429)
                return render(request, 'openalex/landing.html', {
                    'form': form,
                    'error': 'Günlük demo limitiniz doldu.',
                    'remaining': 0,
                    'daily_limit': daily_limit,
                })

            query_parts = form.cleaned_data['query_parts_json']

            job = AlexSearchJob.objects.create(
                user=user,
                query_parts=query_parts,
            )
            run_scraping_job(job.id)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'started',
                    'job_id': str(job.id),
                    'remaining': remaining - 1,
                    'daily_limit': daily_limit,
                })

            return render(request, 'openalex/landing.html', {
                'form': form,
                'job_id': str(job.id),
                'remaining': remaining - 1,
                'daily_limit': daily_limit,
            })
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = form.errors.get('query_parts_json', ['Geçersiz sorgu.'])
                return JsonResponse({'error': errors[0]}, status=400)

    active_job = AlexSearchJob.objects.filter(
        user=user,
        status__in=['pending', 'running'],
    ).order_by('-created_at').first()

    return render(request, 'openalex/landing.html', {
        'form': form,
        'remaining': remaining,
        'daily_limit': daily_limit,
        'active_job_id': str(active_job.id) if active_job else None,
    })


@login_required
@require_GET
def openalex_job_status(request, job_id):
    """AJAX: job durumu + demo sonuçlar."""
    job = get_object_or_404(AlexSearchJob, id=job_id, user=request.user)

    response = {
        'status': job.status,
        'total_results': job.total_results,
    }
    if job.status == 'pending':
        from analizdestek.job_queue import get_queue_position
        response['queue_position'] = get_queue_position('openalex', str(job.id))
    elif job.status == 'completed':
        response['demo_results'] = job.demo_results[:3]
        response['all_results_file_url'] = job.all_results_file_url
    elif job.status == 'failed':
        response['error'] = job.error_message

    return JsonResponse(response)


@login_required
@require_GET
def openalex_download_excel(request, job_id):
    """demo_results JSON'ından Excel (.xlsx) üretir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    job = get_object_or_404(AlexSearchJob, id=job_id, user=request.user)
    records = job.demo_results or []
    if not records:
        raise Http404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OpenAlex Sonuçları'

    headers = ['No', 'Başlık', 'Yazarlar', 'Dergi', 'Yıl', 'DOI', 'Tür', 'Atıf Sayısı', 'Kurumlar', 'Anahtar Kelimeler', 'Açık Erişim', 'Özet']
    ws.append(headers)
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for i, r in enumerate(records, 1):
        kw = r.get('keywords', [])
        ws.append([
            i,
            r.get('title', ''),
            r.get('authors', ''),
            r.get('journal', ''),
            r.get('year', ''),
            r.get('doi', ''),
            r.get('type', ''),
            r.get('cited_by_count', 0),
            r.get('institutions', ''),
            ', '.join(kw) if isinstance(kw, list) else str(kw),
            'Evet' if r.get('open_access') else 'Hayır',
            r.get('abstract', ''),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "xlsx")}"'
    wb.save(response)
    return response


@login_required
@require_POST
def openalex_cancel(request, job_id):
    """Devam eden taramayı iptal eder."""
    job = get_object_or_404(AlexSearchJob, id=job_id, user=request.user)
    if job.status in ('pending', 'running'):
        job.status = 'failed'
        job.error_message = 'Kullanıcı tarafından iptal edildi.'
        job.save(update_fields=['status', 'error_message'])
    return JsonResponse({'success': True})


@login_required
@require_POST
def openalex_send_demo_email(request, job_id):
    """Demo sonuçları kullanıcının kayıtlı emailine gönder."""
    job = get_object_or_404(AlexSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or not job.demo_results:
        return JsonResponse({'error': 'Sonuçlar henüz hazır değil.'}, status=400)

    if job.demo_email_sent:
        return JsonResponse({'error': 'Demo sonuçlar zaten gönderildi.'}, status=400)

    from .services.job_runner import send_demo_email_async
    send_demo_email_async(job.id)

    return JsonResponse({'success': True, 'message': f'Demo sonuçların {request.user.email} adresine gönderilmesi için işlem başlatıldı.'})


@login_required
def openalex_order_page(request, job_id):
    """Sipariş sayfası: GET=form göster, POST=sipariş oluştur."""
    job = get_object_or_404(AlexSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or job.total_results == 0:
        return render(request, 'openalex/order.html', {
            'job': job,
            'error': 'Bu arama için henüz sonuç yok.',
        })

    existing_order = AlexOrder.objects.filter(search_job=job, user=request.user).first()
    if existing_order:
        return render(request, 'openalex/order.html', {
            'job': job,
            'existing_order': existing_order,
        })

    if request.method == 'POST':
        form = AlexOrderForm(request.POST)
        if form.is_valid():
            abstract_count = min(form.cleaned_data['abstract_count'], job.total_results)
            price = AlexOrder.calculate_price(abstract_count)

            order = AlexOrder.objects.create(
                user=request.user,
                search_job=job,
                abstract_count=abstract_count,
                total_price=price,
                payment_note=form.cleaned_data.get('payment_note', ''),
                status='pending_payment',
            )
            return render(request, 'openalex/order.html', {
                'job': job,
                'existing_order': order,
                'success': True,
            })

    return render(request, 'openalex/order.html', {
        'job': job,
    })
