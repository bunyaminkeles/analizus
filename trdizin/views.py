from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, Http404, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET, require_POST
from functools import wraps

from .forms import DizinSearchForm, DizinOrderForm
from .models import DizinSearchJob, DizinOrder
from .services.job_runner import run_scraping_job
from tarama_seo_content import TARAMA_SEO_CONTENT


def _safe_filename(job, ext: str) -> str:
    """trdizin_<arama_kelimesi>_<YYYYMMDD>.<ext> formatında güvenli dosya adı."""
    import re
    from django.utils import timezone
    keyword = ''
    if job.query_parts:
        for part in job.query_parts:
            val = part.get('value', '').strip()
            if val:
                keyword = val
                break
    keyword = (keyword or 'sonuclar')[:40]
    keyword = re.sub(r'[^\w\s-]', '', keyword, flags=re.UNICODE)
    keyword = re.sub(r'[\s]+', '_', keyword).strip('_')
    date_str = timezone.now().strftime('%Y%m%d')
    return f'trdizin_{keyword}_{date_str}.{ext}'


def feature_required(flag_name):
    """Decorator: SiteSettings'deki feature flag kapalıysa 404 döner."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            from forum.models import SiteSettings
            site = SiteSettings.load()
            if not getattr(site, f'feature_{flag_name}', True):
                raise Http404
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


IBAN_INFO = {
    'hesap_sahibi': 'Bünyamin Keleş',
    'iban': 'TR73 0003 2000 0000 0079 1034 65',
}


@feature_required('trdizin')
def trdizin_landing(request):
    """Landing page: gelişmiş arama formu + demo arama."""
    if not request.user.is_authenticated:
        return render(request, 'service_promo.html', {
            'promo_title': 'TR Dizin Makale Kazıma ve İndirme Aracı',
            'promo_icon': 'bi-journal-text',
            'promo_color': 'primary',
            'promo_description': 'TR Dizin\'den kodsuz makale veri kazıma aracı. Anahtar kelime, yazar ve dergi bazında gelişmiş arama yapın; sonuçları Excel veya TXT olarak indirin. Python bilgisi gerekmez.',
            'promo_features': [
                {'icon': 'bi-download', 'title': 'Excel & TXT İndirme', 'desc': 'Makale başlığı, yazar, dergi, yıl ve özet verilerini tek tıkla Excel veya TXT olarak indirin.'},
                {'icon': 'bi-search', 'title': 'Kodsuz Veri Kazıma', 'desc': 'Başlık, yazar, dergi, anahtar kelime ve özet alanlarını birleştirerek binlerce makaleden veri çekin.'},
                {'icon': 'bi-file-earmark-text-fill', 'title': 'Tam Künyeye Erişim', 'desc': 'TÜBİTAK ULAKBİM hakemli dergilerindeki makalelerin DOI, özet ve atıf bilgilerine erişin.'},
                {'icon': 'bi-bar-chart-fill', 'color': 'warning', 'title': 'Tek Tıkla Analiz', 'desc': '10+ sonuçta Analiz Yap butonu ile yazar, dergi ve yıl dağılım grafiklerini PDF olarak alın.'},
            ],
            'promo_steps': [
                'Arama alanlarına anahtar kelime, yazar adı veya dergi adını girin.',
                'TR Dizin\'den makaleler saniyeler içinde kazınır ve listelenir.',
                'Excel veya TXT olarak indirin ya da Analiz Yap ile bibliometrik rapor alın.',
            ],
        })

    form = DizinSearchForm()
    user = request.user

    daily_count = DizinSearchJob.daily_count_for_user(user)
    daily_limit = DizinSearchJob.get_daily_limit(user)
    remaining = max(0, daily_limit - daily_count)

    if request.method == 'POST':
        form = DizinSearchForm(request.POST)
        if form.is_valid():
            if not user.profile.email_verified:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'error': 'TR Dizin tarama için e-posta doğrulaması gereklidir. Profil sayfanızdan e-postanızı doğrulayın.'
                    }, status=403)
                return render(request, 'trdizin/landing.html', {
                    'form': form,
                    'error': 'TR Dizin tarama için e-posta doğrulaması gereklidir.',
                    'remaining': remaining,
                    'daily_limit': daily_limit,
                })

            if remaining <= 0:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'error': f'Günlük demo limitiniz doldu ({daily_limit}/{daily_limit}). '
                                 f'{"Premium üyelikle 7 aramaya yükseltebilirsiniz." if not user.profile.is_premium else "Yarın tekrar deneyebilirsiniz."}'
                    }, status=429)
                return render(request, 'trdizin/landing.html', {
                    'form': form,
                    'error': 'Günlük demo limitiniz doldu.',
                    'remaining': 0,
                    'daily_limit': daily_limit,
                })

            query_parts = form.cleaned_data['query_parts_json']

            job = DizinSearchJob.objects.create(
                user=user,
                query_parts=query_parts,
            )
            run_scraping_job(job.id)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'started',
                    'job_id': str(job.id),
                    'remaining': remaining - 1,
                    'daily_limit': daily_limit,
                })

            return render(request, 'trdizin/landing.html', {
                'form': form,
                'job_id': str(job.id),
                'remaining': remaining - 1,
                'daily_limit': daily_limit,
            })
        else:
            # Form validation hatası
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = form.errors.get('query_parts_json', ['Geçersiz sorgu.'])
                return JsonResponse({'error': errors[0]}, status=400)

    active_job = DizinSearchJob.objects.filter(
        user=user,
        status__in=['pending', 'running'],
    ).order_by('-created_at').first()

    return render(request, 'trdizin/landing.html', {
        'form': form,
        'remaining': remaining,
        'daily_limit': daily_limit,
        'active_job_id': str(active_job.id) if active_job else None,
        'seo_guide': TARAMA_SEO_CONTENT.get('trdizin'),
    })


@login_required
@require_GET
def trdizin_job_status(request, job_id):
    """AJAX: job durumu + demo sonuçlar."""
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)

    response = {
        'status': job.status,
        'total_results': job.total_results,
    }
    if job.status == 'pending':
        from analizdestek.job_queue import get_queue_position
        response['queue_position'] = get_queue_position('trdizin', str(job.id))
    elif job.status == 'completed':
        response['demo_results'] = job.demo_results[:5]
        response['all_results_file_url'] = job.all_results_file_url
    elif job.status == 'failed':
        response['error'] = job.error_message

    return JsonResponse(response)


@login_required
@require_POST
def trdizin_send_demo_email(request, job_id):
    """Demo sonuçları kullanıcının kayıtlı emailine gönder."""
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or not job.demo_results:
        return JsonResponse({'error': 'Sonuçlar henüz hazır değil.'}, status=400)

    if job.demo_email_sent:
        return JsonResponse({'error': 'Demo sonuçlar zaten gönderildi.'}, status=400)

    from .services.job_runner import send_demo_email_async
    send_demo_email_async(job.id)

    return JsonResponse({'success': True, 'message': f'Demo sonuçların {request.user.email} adresine gönderilmesi için işlem başlatıldı.'})


@login_required
def trdizin_order_page(request, job_id):
    """Sipariş sayfası: GET=form göster, POST=sipariş oluştur."""
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or job.total_results == 0:
        return render(request, 'trdizin/order.html', {
            'job': job,
            'error': 'Bu arama için henüz sonuç yok.',
        })

    # Zaten sipariş var mı?
    existing_order = DizinOrder.objects.filter(search_job=job, user=request.user).first()
    if existing_order:
        return render(request, 'trdizin/order.html', {
            'job': job,
            'existing_order': existing_order,
        })

    if request.method == 'POST':
        form = DizinOrderForm(request.POST)
        if form.is_valid():
            abstract_count = min(form.cleaned_data['abstract_count'], job.total_results)
            price = DizinOrder.calculate_price(abstract_count)

            order = DizinOrder.objects.create(
                user=request.user,
                search_job=job,
                abstract_count=abstract_count,
                total_price=price,
                payment_note=form.cleaned_data.get('payment_note', ''),
                status='pending_payment',
            )
            return render(request, 'trdizin/order.html', {
                'job': job,
                'existing_order': order,
                'success': True,
            })

    return render(request, 'trdizin/order.html', {
        'job': job,
    })


@login_required
@require_GET
def trdizin_download_excel(request, job_id):
    """demo_results JSON'ından Excel (.xlsx) üretir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)
    records = job.demo_results or []
    if not records:
        raise Http404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TR Dizin Sonuçları'

    headers = ['No', 'Başlık', 'Yazarlar', 'Dergi', 'Yıl', 'DOI', 'Yayın Türü', 'Erişim', 'Dil', 'Anahtar Kelimeler (TR)', 'Özet (TR)', 'Özet (EN)']
    ws.append(headers)
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for i, r in enumerate(records, 1):
        kw_tr = r.get('keywords_tr', [])
        ws.append([
            i,
            r.get('title', ''),
            r.get('authors', ''),
            r.get('journal', ''),
            r.get('year', ''),
            r.get('doi', ''),
            r.get('publication_type', ''),
            r.get('access_type', ''),
            r.get('language', ''),
            ', '.join(kw_tr) if isinstance(kw_tr, list) else str(kw_tr),
            r.get('abstract_tr', ''),
            r.get('abstract_en', ''),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "xlsx")}"'
    wb.save(response)
    return response


@login_required
@require_GET
def trdizin_download(request, job_id):
    """S3'teki TXT dosyasını proxy ile indirtir (tarayıcıda açmaz)."""
    import requests as req_lib
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)

    if not job.all_results_file_url:
        raise Http404

    try:
        s3_resp = req_lib.get(job.all_results_file_url, timeout=30, stream=True)
        s3_resp.raise_for_status()
    except Exception:
        raise Http404

    response = HttpResponse(
        s3_resp.content,
        content_type='text/plain; charset=utf-8',
    )
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "txt")}"'
    return response


@login_required
@require_POST
def trdizin_cancel(request, job_id):
    """Devam eden taramayı iptal eder."""
    job = get_object_or_404(DizinSearchJob, id=job_id, user=request.user)
    if job.status in ('pending', 'running'):
        job.status = 'failed'
        job.error_message = 'Kullanıcı tarafından iptal edildi.'
        job.save(update_fields=['status', 'error_message'])
    return JsonResponse({'success': True})
