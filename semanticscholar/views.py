import re
import logging
from functools import wraps

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_POST, require_GET
from django_ratelimit.decorators import ratelimit

from .forms import SemanticSearchForm, SemanticOrderForm
from .models import SemanticSearchJob, SemanticOrder
from .services.job_runner import run_scraping_job
from tarama_seo_content import TARAMA_SEO_CONTENT

logger = logging.getLogger(__name__)


def feature_required(flag_name):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            from forum.models import SiteSettings
            if not getattr(SiteSettings.load(), f'feature_{flag_name}', False):
                raise Http404
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator


def _safe_filename(job, ext):
    keyword = ''
    if job.query_parts:
        for part in job.query_parts:
            val = part.get('value', '').strip()
            if val:
                keyword = val
                break
    from django.utils import timezone
    keyword = (keyword or 'sonuclar')[:40]
    keyword = re.sub(r'[^\w\s-]', '', keyword, flags=re.UNICODE)
    keyword = re.sub(r'\s+', '_', keyword).strip('_')
    return f"semantic_scholar_{keyword}_{timezone.now().strftime('%Y%m%d')}.{ext}"


IBAN_INFO = {
    'hesap_sahibi': 'Bünyamin Keleş',
    'iban': 'TR73 0003 2000 0000 0079 1034 65',
}


@feature_required('semanticscholar')
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def semantic_landing(request):
    if not request.user.is_authenticated:
        return render(request, 'service_promo.html', {
            'promo_title': 'Semantic Scholar Yayın Kazıma',
            'promo_icon': 'bi-diagram-3-fill',
            'promo_color': 'info',
            'promo_description': '200 milyondan fazla akademik yayın arasında arama yapın. Başlık, yazar, alan ve yıl filtresiyle arama sonuçlarını Excel veya TXT olarak indirin. Python bilgisi gerekmez.',
            'promo_features': [
                {'icon': 'bi-search', 'title': 'Geniş Kapsam', 'desc': '200M+ yayın — WoS ve Scopus dahil tüm büyük veri tabanlarını kapsıyor.'},
                {'icon': 'bi-download', 'title': 'Excel & TXT İndirme', 'desc': 'Başlık, yazar, dergi, yıl, atıf, özet ve DOI verilerini tek tıkla indirin.'},
                {'icon': 'bi-link-45deg', 'title': 'CrossRef Zenginleştirme', 'desc': 'DOI\'si olan kayıtlara kurum, yayıncı ve konu bilgisi otomatik eklenir.'},
                {'icon': 'bi-unlock-fill', 'title': 'Açık Erişim', 'desc': 'Açık erişimli yayınlar için PDF linki de gösterilir.'},
            ],
            'promo_steps': [
                'Arama kutusuna anahtar kelime, yazar veya araştırma alanı girin.',
                'Sistem Semantic Scholar\'dan yayınları otomatik çeker.',
                'Demo 5 sonuç ücretsiz; tamamı için sipariş oluşturun.',
            ],
            'seo_guide': TARAMA_SEO_CONTENT.get('semanticscholar'),
        })

    user = request.user

    if hasattr(user, 'profile') and not user.profile.email_verified:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'Semantic Scholar tarama için e-posta adresinizi doğrulamanız gerekiyor.')
        return redirect('profile_edit')

    daily_limit = SemanticSearchJob.get_daily_limit(user)
    daily_used = SemanticSearchJob.daily_count_for_user(user)
    remaining = max(0, daily_limit - daily_used)

    if request.method == 'POST':
        form = SemanticSearchForm(request.POST)
        if form.is_valid():
            if remaining <= 0:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'error': f'Günlük demo limitiniz doldu ({daily_limit}/{daily_limit}). Yarın tekrar deneyebilirsiniz.'
                    }, status=429)
                return render(request, 'semanticscholar/landing.html', {
                    'form': form,
                    'error': 'Günlük demo limitiniz doldu.',
                    'remaining': 0,
                    'daily_limit': daily_limit,
                })

            query_parts = form.cleaned_data['query_parts_json']
            job = SemanticSearchJob.objects.create(user=user, query_parts=query_parts)
            run_scraping_job(job.id)

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'started',
                    'job_id': str(job.id),
                    'remaining': remaining - 1,
                    'daily_limit': daily_limit,
                })

            return render(request, 'semanticscholar/landing.html', {
                'form': form,
                'job_id': str(job.id),
                'remaining': remaining - 1,
                'daily_limit': daily_limit,
            })
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = form.errors.get('query_parts_json', ['Geçersiz sorgu.'])
                return JsonResponse({'error': errors[0]}, status=400)

    form = SemanticSearchForm()
    active_job = SemanticSearchJob.objects.filter(
        user=user, status__in=['pending', 'running']
    ).order_by('-created_at').first()

    return render(request, 'semanticscholar/landing.html', {
        'form': form,
        'remaining': remaining,
        'daily_limit': daily_limit,
        'active_job_id': str(active_job.id) if active_job else None,
        'seo_guide': TARAMA_SEO_CONTENT.get('semanticscholar'),
    })


@login_required
@require_GET
def semantic_job_status(request, job_id):
    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)
    response = {
        'status': job.status,
        'total_results': job.total_results,
    }
    if job.status == 'pending':
        from analizdestek.job_queue import get_queue_position
        response['queue_position'] = get_queue_position('semanticscholar', str(job.id))
    elif job.status == 'completed':
        response['demo_results'] = job.demo_results[:5]
        response['all_results_file_url'] = job.all_results_file_url
    elif job.status == 'failed':
        response['error'] = job.error_message
    return JsonResponse(response)


@login_required
@require_POST
def semantic_send_demo_email(request, job_id):
    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)
    if job.status != 'completed' or not job.demo_results:
        return JsonResponse({'error': 'Sonuçlar henüz hazır değil.'}, status=400)
    if job.demo_email_sent:
        return JsonResponse({'error': 'Demo sonuçlar zaten gönderildi.'}, status=400)
    from .services.job_runner import send_demo_email_async
    send_demo_email_async(job.id)
    return JsonResponse({'success': True, 'message': f'Demo sonuçların {request.user.email} adresine gönderilmesi başlatıldı.'})


@login_required
def semantic_order_page(request, job_id):
    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or job.total_results == 0:
        return render(request, 'semanticscholar/order.html', {
            'job': job,
            'error': 'Bu arama için henüz sonuç yok.',
        })

    existing_order = SemanticOrder.objects.filter(search_job=job, user=request.user).first()
    if existing_order:
        return render(request, 'semanticscholar/order.html', {
            'job': job,
            'existing_order': existing_order,
            'iban_info': IBAN_INFO,
        })

    if request.method == 'POST':
        form = SemanticOrderForm(request.POST)
        if form.is_valid():
            abstract_count = min(form.cleaned_data['abstract_count'], job.total_results)
            price = SemanticOrder.calculate_price(abstract_count)
            order = SemanticOrder.objects.create(
                user=request.user,
                search_job=job,
                abstract_count=abstract_count,
                total_price=price,
                payment_note=form.cleaned_data.get('payment_note', ''),
                status='pending_payment',
            )
            return render(request, 'semanticscholar/order.html', {
                'job': job,
                'existing_order': order,
                'success': True,
                'iban_info': IBAN_INFO,
            })

    return render(request, 'semanticscholar/order.html', {
        'job': job,
        'form': SemanticOrderForm(),
        'iban_info': IBAN_INFO,
    })


@login_required
@require_GET
def semantic_download_excel(request, job_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)
    records = job.demo_results or []
    if not records:
        raise Http404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Semantic Scholar Sonuçları'

    headers = ['No', 'Başlık', 'Yazarlar', 'Dergi/Kaynak', 'Yıl', 'DOI', 'Yayın Türü',
               'Atıf Sayısı', 'Kurumlar', 'Araştırma Alanları', 'OA PDF', 'Özet']
    ws.append(headers)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for i, r in enumerate(records, 1):
        fos = r.get('fields_of_study', [])
        ws.append([
            i,
            r.get('title', ''),
            r.get('authors', ''),
            r.get('journal', ''),
            r.get('year', ''),
            r.get('doi', ''),
            r.get('type', ''),
            r.get('cited_by_count', 0),
            r.get('institutions', ''),
            '; '.join(fos) if isinstance(fos, list) else str(fos),
            r.get('open_access_pdf', ''),
            r.get('abstract', ''),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "xlsx")}"'
    wb.save(response)
    return response


@login_required
@require_GET
def semantic_download(request, job_id):
    import requests as req_lib
    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)
    if not job.all_results_file_url:
        raise Http404
    try:
        s3_resp = req_lib.get(job.all_results_file_url, timeout=30, stream=True)
        s3_resp.raise_for_status()
    except Exception:
        raise Http404
    response = HttpResponse(s3_resp.content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "txt")}"'
    return response


@login_required
@require_POST
def semantic_cancel(request, job_id):
    job = get_object_or_404(SemanticSearchJob, id=job_id, user=request.user)
    if job.status in ('pending', 'running'):
        job.status = 'failed'
        job.error_message = 'Kullanıcı tarafından iptal edildi.'
        job.save(update_fields=['status', 'error_message'])
    return JsonResponse({'success': True})
