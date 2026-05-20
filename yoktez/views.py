import logging
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET, require_POST
from django.http import Http404
from django_ratelimit.decorators import ratelimit
from functools import wraps

from .forms import YokTezSearchForm
from .models import YokTezSearchJob

logger = logging.getLogger(__name__)


def _safe_filename(job, ext: str) -> str:
    """yoktez_<arama_kelimesi>_<YYYYMMDD>.<ext> formatında güvenli dosya adı üretir."""
    import re
    from django.utils import timezone
    keyword = job.tez_ad or job.yazar or job.metin or job.danisman or job.universite or 'sonuclar'
    keyword = keyword[:40].strip()
    keyword = re.sub(r'[^\w\s-]', '', keyword, flags=re.UNICODE)
    keyword = re.sub(r'[\s]+', '_', keyword).strip('_')
    date_str = timezone.now().strftime('%Y%m%d')
    return f'yoktez_{keyword}_{date_str}.{ext}'


def feature_required(flag_name):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            from forum.models import SiteSettings
            site = SiteSettings.load()
            if not getattr(site, f'feature_{flag_name}', True):
                raise Http404
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


@feature_required('yoktez')
@ratelimit(key='user', rate='10/h', method='POST', block=True)
def yoktez_landing(request):
    if not request.user.is_authenticated:
        return render(request, 'service_promo.html', {
            'promo_title': 'YÖK Tez Kazıma ve İndirme Aracı',
            'promo_icon': 'bi-mortarboard-fill',
            'promo_color': 'success',
            'promo_description': 'YÖK Ulusal Tez Merkezi\'nden kodsuz veri kazıma aracı. Anahtar kelime, yazar veya danışmanla arama yapın; tez verilerini Excel veya TXT olarak indirin. Python bilgisi gerekmez.',
            'promo_features': [
                {'icon': 'bi-download', 'title': 'Excel & TXT İndirme', 'desc': 'Tez No, Başlık, Yazar, Danışman, Üniversite, Yıl ve Özet verilerini tek tıkla Excel veya TXT olarak indirin.'},
                {'icon': 'bi-search', 'title': 'Kodsuz Veri Kazıma', 'desc': 'Python veya Selenium bilgisi gerekmeden, anahtar kelime ile binlerce tez verisini saniyeler içinde çekin.'},
                {'icon': 'bi-person-lines-fill', 'title': 'Danışman & Kurum Verisi', 'desc': 'Danışman adı, üniversite ve yıl bilgisini yapılandırılmış biçimde edinin.'},
                {'icon': 'bi-graph-up-arrow', 'color': 'warning', 'title': 'Tek Tıkla Bibliometrik Analiz', 'desc': '10+ sonuçta Analiz Yap butonu ile trend ve dağılım grafiklerini PDF olarak alın.'},
            ],
            'promo_steps': [
                'Arama kutusuna anahtar kelime, konu veya yazar adı girin.',
                'YÖK Tez Merkezi\'nden veriler saniyeler içinde kazınır ve listelenir.',
                'Excel veya TXT olarak indirin ya da Analiz Yap ile bibliometrik rapor alın.',
            ],
        })
    user = request.user
    daily_count = YokTezSearchJob.daily_count_for_user(user)
    daily_limit = YokTezSearchJob.get_daily_limit(user)
    remaining = max(0, daily_limit - daily_count)

    if request.method == 'POST':
        form = YokTezSearchForm(request.POST)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if not user.profile.email_verified:
            if is_ajax:
                return JsonResponse({'error': 'E-posta doğrulaması gereklidir.'}, status=403)
            return render(request, 'yoktez/landing.html', {
                'form': form, 'remaining': remaining, 'daily_limit': daily_limit,
                'error': 'E-posta doğrulaması gereklidir.',
            })

        if remaining <= 0:
            msg = (f'Günlük arama limitiniz doldu ({daily_limit}/{daily_limit}). '
                   f'{"Premium üyelikle 7 aramaya yükseltebilirsiniz." if not user.profile.is_premium else "Yarın tekrar deneyebilirsiniz."}')
            if is_ajax:
                return JsonResponse({'error': msg}, status=429)
            return render(request, 'yoktez/landing.html', {
                'form': form, 'remaining': 0, 'daily_limit': daily_limit, 'error': msg,
            })

        if form.is_valid():
            cd = form.cleaned_data
            job = YokTezSearchJob.objects.create(
                user=user,
                tez_ad=cd.get('tez_ad', ''),
                yazar=cd.get('yazar', ''),
                danisman=cd.get('danisman', ''),
                universite=cd.get('universite', ''),
                tur=cd.get('tur', '0'),
                yil_baslangic=cd.get('yil_baslangic'),
                yil_bitis=cd.get('yil_bitis'),
                metin=cd.get('metin', ''),
            )

            from .services.job_runner import run_yoktez_job
            run_yoktez_job(str(job.id))

            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'job_id': str(job.id),
                    'remaining': remaining - 1,
                })
        else:
            if is_ajax:
                errors = [e for field_errors in form.errors.values() for e in field_errors]
                return JsonResponse({'error': errors[0] if errors else 'Geçersiz form.'}, status=400)

    else:
        form = YokTezSearchForm()

    active_job = YokTezSearchJob.objects.filter(
        user=user,
        status__in=['pending', 'running'],
    ).order_by('-created_at').first()

    # Tamamlanmış son job — kullanıcı sayfadan ayrılıp dönünce sonuçları göster
    completed_job = None
    if not active_job and request.method == 'GET':
        from django.utils import timezone
        cutoff = timezone.now() - timezone.timedelta(hours=24)
        completed_job = YokTezSearchJob.objects.filter(
            user=user,
            status='completed',
            completed_at__gte=cutoff,
        ).order_by('-completed_at').first()

    restore_job = active_job or completed_job
    if restore_job and request.method == 'GET':
        form = YokTezSearchForm(initial={
            'tez_ad': restore_job.tez_ad or '',
            'universite': restore_job.universite or '',
            'tur': restore_job.tur or '0',
            'yil_baslangic': restore_job.yil_baslangic,
            'yil_bitis': restore_job.yil_bitis,
            'metin': restore_job.metin or '',
        })

    from tezanaliz.models import TezAnaliz
    past_analiz_jobs = TezAnaliz.objects.filter(user=user).select_related('yok_job').order_by('-created_at')[:8]

    return render(request, 'yoktez/landing.html', {
        'form': form,
        'remaining': remaining,
        'daily_limit': daily_limit,
        'active_job_id': str(active_job.id) if active_job else None,
        'completed_job_id': str(completed_job.id) if completed_job else None,
        'past_analiz_jobs': past_analiz_jobs,
    })


@login_required
@require_GET
def yoktez_job_status(request, job_id):
    from django.shortcuts import get_object_or_404
    job = get_object_or_404(YokTezSearchJob, id=job_id, user=request.user)

    resp = {'status': job.status, 'total_results': job.total_results}
    if job.status == 'pending':
        from analizdestek.job_queue import get_queue_position
        resp['queue_position'] = get_queue_position('yoktez', str(job.id))
    elif job.status == 'completed':
        resp['demo_results'] = job.demo_results[:5]
        resp['all_results_file_url'] = job.all_results_file_url
    elif job.status == 'failed':
        resp['error'] = job.error_message
    return JsonResponse(resp)


@login_required
@login_required
@require_GET
def yoktez_download_excel(request, job_id):
    """demo_results JSON'ından Excel (.xlsx) üretir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    job = get_object_or_404(YokTezSearchJob, id=job_id, user=request.user)
    records = job.demo_results or []
    if not records:
        raise Http404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'YÖK Tez Sonuçları'

    headers = ['No', 'Tez No', 'Başlık (EN)', 'Başlık (TR)', 'Yazar', 'Danışman', 'Üniversite', 'Yıl', 'Tür', 'Dil', 'Özet (TR)', 'Özet (EN)']
    ws.append(headers)
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for i, r in enumerate(records, 1):
        ws.append([
            i,
            r.get('tez_no', ''),
            r.get('title', ''),
            r.get('title_tr', ''),
            r.get('author', ''),
            r.get('danisman', ''),
            r.get('university', ''),
            r.get('year', ''),
            r.get('thesis_type', ''),
            r.get('language', ''),
            r.get('abstract_tr', ''),
            r.get('abstract_en', ''),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "xlsx")}"'
    wb.save(response)
    return response


@login_required
@require_GET
def yoktez_download(request, job_id):
    """S3'teki TXT dosyasını proxy ile indirtir; S3 yoksa demo_results'tan üretir."""
    import requests as req_lib
    job = get_object_or_404(YokTezSearchJob, id=job_id, user=request.user)

    if job.all_results_file_url:
        try:
            s3_resp = req_lib.get(job.all_results_file_url, timeout=30, stream=True)
            s3_resp.raise_for_status()
            response = HttpResponse(
                s3_resp.content,
                content_type='text/plain; charset=utf-8',
            )
            response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "txt")}"'
            return response
        except Exception:
            pass

    # Fallback: demo_results'tan on-the-fly üret
    if not job.demo_results:
        raise Http404

    from .services.scraper import generate_results_txt
    txt = generate_results_txt(job.demo_results, job)
    response = HttpResponse(txt, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{_safe_filename(job, "txt")}"'
    return response


@login_required
@require_POST
def yoktez_send_demo_email(request, job_id):
    from django.shortcuts import get_object_or_404
    job = get_object_or_404(YokTezSearchJob, id=job_id, user=request.user)

    if job.status not in ('completed',):
        return JsonResponse({'error': 'Sonuçlar henüz hazır değil.'}, status=400)
    if job.total_results == 0:
        return JsonResponse({'error': 'Arama sonucu bulunamadı, gönderilecek bir şey yok.'}, status=400)
    if job.demo_email_sent:
        return JsonResponse({'error': 'Demo sonuçlar zaten gönderildi.'}, status=400)

    from .services.job_runner import send_demo_email_async
    send_demo_email_async(str(job.id))
    return JsonResponse({'success': True, 'message': f'Sonuçlar {request.user.email} adresine gönderildi.'})


@login_required
@require_POST
def yoktez_cancel(request, job_id):
    """Devam eden taramayı iptal eder."""
    job = get_object_or_404(YokTezSearchJob, id=job_id, user=request.user)
    if job.status in ('pending', 'running'):
        job.status = 'failed'
        job.error_message = 'Kullanıcı tarafından iptal edildi.'
        job.save(update_fields=['status', 'error_message'])
    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Tez Analizi view'ları (tezanaliz app'ten taşındı — /yoktez/analiz/... URL'leri)
# ---------------------------------------------------------------------------

@login_required
@require_POST
def create_analiz(request, yok_job_id):
    """YÖK Tez aramasından tez analizi başlat."""
    from tezanaliz.models import TezAnaliz

    try:
        yok_job = YokTezSearchJob.objects.get(id=yok_job_id, user=request.user)
    except YokTezSearchJob.DoesNotExist:
        return JsonResponse({'error': 'YÖK Tez araması bulunamadı.'}, status=404)

    if yok_job.total_results < 10:
        return JsonResponse(
            {'error': f'Analiz için en az 10 sonuç gereklidir (bulunan: {yok_job.total_results}).'},
            status=400,
        )

    limit = TezAnaliz.get_daily_limit(request.user)
    used = TezAnaliz.daily_count_for_user(request.user)
    if used >= limit:
        return JsonResponse({'error': f'Günlük analiz limitinize ({limit}) ulaştınız.'}, status=429)

    existing = TezAnaliz.objects.filter(yok_job=yok_job, user=request.user).first()
    if existing and existing.status in ('pending', 'running'):
        return JsonResponse({'job_id': str(existing.id), 'already_running': True})
    if existing and existing.status == 'completed':
        return JsonResponse({'job_id': str(existing.id), 'already_completed': True})

    job = TezAnaliz.objects.create(
        user=request.user,
        yok_job=yok_job,
        tez_ad=yok_job.tez_ad,
        yazar=yok_job.yazar,
        universite=yok_job.universite,
        tur=yok_job.tur,
        yil_baslangic=yok_job.yil_baslangic,
        yil_bitis=yok_job.yil_bitis,
        metin=yok_job.metin,
    )

    from tezanaliz.services.job_runner import run_tezanaliz_job
    run_tezanaliz_job(str(job.id))

    return JsonResponse({'job_id': str(job.id)})


STALE_TIMEOUT_MINUTES = 15


@login_required
def analiz_status(request, job_id):
    """Analiz iş durumunu döndür."""
    from django.utils import timezone
    from tezanaliz.models import TezAnaliz

    job = get_object_or_404(TezAnaliz, id=job_id, user=request.user)

    if job.status in ('pending', 'running'):
        elapsed = (timezone.now() - job.created_at).total_seconds()
        if elapsed > STALE_TIMEOUT_MINUTES * 60:
            job.mark_failed(
                f'İşlem {STALE_TIMEOUT_MINUTES} dakika içinde tamamlanamadı. '
                'YÖK Tez sunucusu yavaş yanıt veriyor olabilir. Lütfen tekrar deneyin.'
            )

    queue_position = 0
    if job.status == 'pending':
        from analizdestek.job_queue import get_queue_position
        queue_position = get_queue_position('tezanaliz', str(job.id))

    return JsonResponse({
        'status': job.status,
        'total_records': job.total_records,
        'pdf_url': job.pdf_url,
        'error': job.error_message,
        'queue_position': queue_position,
    })


@login_required
def analiz_results(request, job_id):
    """Analiz sonuç sayfası."""
    from tezanaliz.models import TezAnaliz

    job = get_object_or_404(TezAnaliz, id=job_id, user=request.user)
    similar = (job.analysis_data or {}).get('similar', [])

    analyses_list = [
        'Tez Türlerine Göre Dağılım (Pasta Grafik)',
        'Yıllara Göre Tez Türü Trendi (Stacked Bar)',
        'Üniversite Bazında Üretkenlik (Top 15)',
        'TF-IDF Anahtar Kelimeler (Başlık + Özet)',
        'Son 5 Yıl Tez Trendi + Büyüme Oranı',
        'LDA Konu Modelleme (Gizli Konular)',
        'Konu & Dizin Terimleri Kelime Bulutu',
    ]
    return render(request, 'yoktez/results.html', {
        'job': job,
        'similar': similar,
        'analyses_list': analyses_list,
    })
