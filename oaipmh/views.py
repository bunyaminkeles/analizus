import json
import logging
from datetime import timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from django_ratelimit.decorators import ratelimit

from forum.models import SiteSettings
from .models import University, OAIPMHSearchJob, OAIPMHOrder
from .forms import OAIPMHKeywordForm, OAIPMHBrowseForm, OAIPMHOrderForm
from .services.job_runner import run_scraping_job, send_demo_email_async
from tarama_seo_content import TARAMA_SEO_CONTENT

logger = logging.getLogger(__name__)

IBAN_INFO = {
    'hesap_sahibi': 'Bünyamin Keleş',
    'iban': 'TR73 0003 2000 0000 0079 1034 65',
}


def _feature_check():
    site = SiteSettings.load()
    if not site.feature_oaipmh:
        raise Http404


def _check_email_verified(request):
    profile = getattr(request.user, 'profile', None)
    if profile and not profile.email_verified:
        return False
    return True


@ratelimit(key='user', rate='10/h', method='POST', block=True)
def oaipmh_landing(request):
    _feature_check()

    if not request.user.is_authenticated:
        return render(request, 'service_promo.html', {
            'promo_title': 'Üniversite Tez Arşivi',
            'promo_icon': 'bi-mortarboard',
            'promo_color': 'warning',
            'promo_description': '19 Türk üniversitesinin açık erişim arşivinde anahtar kelimeyle arama yapın veya bir üniversitenin tüm tezlerini tarayın. ODTÜ, İTÜ, Sabancı dahil 19 üniversite.',
            'promo_features': [
                {'icon': 'bi-building-fill', 'color': 'warning', 'title': '19 Üniversite', 'desc': 'ODTÜ, İTÜ, Sabancı, Dokuz Eylül, Akdeniz ve daha fazlası.'},
                {'icon': 'bi-unlock-fill', 'color': 'success', 'title': 'Açık Erişim', 'desc': 'OAI-PMH protokolü ile üniversitelerin kendi açık arşivlerinden doğrudan veri.'},
                {'icon': 'bi-search', 'title': 'İki Mod', 'desc': 'Anahtar kelime araması veya üniversite bazlı toplu tarama seçenekleri.'},
                {'icon': 'bi-download', 'title': 'Sonuçları İndir', 'desc': 'Arama sonuçlarını TXT olarak indirin, istediğiniz araca aktarın.'},
            ],
            'promo_steps': [
                'Anahtar kelime girin veya taramak istediğiniz üniversiteyi seçin.',
                'Sistem üniversitelerin açık arşivlerinden tezleri toplar.',
                'Sonuçları indirin veya detaylı inceleyin.',
            ],
        })

    if not _check_email_verified(request):
        return render(request, 'oaipmh/landing.html', {
            'email_not_verified': True,
            'keyword_form': OAIPMHKeywordForm(),
            'browse_form': OAIPMHBrowseForm(),
        })

    daily_used = OAIPMHSearchJob.daily_count_for_user(request.user)
    daily_limit = OAIPMHSearchJob.get_daily_limit(request.user)
    remaining = max(0, daily_limit - daily_used)

    if request.method == 'POST':
        search_type = request.POST.get('search_type', 'keyword')

        if remaining <= 0:
            return JsonResponse({'success': False, 'error': 'Günlük arama limitinize ulaştınız.'}, status=429)

        if search_type == 'keyword':
            form = OAIPMHKeywordForm(request.POST)
            if not form.is_valid():
                errors = '; '.join([str(e) for errs in form.errors.values() for e in errs])
                return JsonResponse({'success': False, 'error': errors}, status=400)
            selected_unis = form.cleaned_data.get('universities')
            job = OAIPMHSearchJob.objects.create(
                user=request.user,
                search_type='keyword',
                keyword=form.cleaned_data.get('keyword', ''),
                abstract_query=form.cleaned_data.get('abstract_query', ''),
                university_ids=[u.id for u in selected_unis] if selected_unis else [],
                year_from=form.cleaned_data.get('year_from'),
                year_to=form.cleaned_data.get('year_to'),
            )
        else:  # browse
            form = OAIPMHBrowseForm(request.POST)
            if not form.is_valid():
                return JsonResponse({'success': False, 'error': str(form.errors)}, status=400)
            job = OAIPMHSearchJob.objects.create(
                user=request.user,
                search_type='browse',
                university=form.cleaned_data['university'],
            )

        run_scraping_job(job.id)
        return JsonResponse({'success': True, 'job_id': str(job.id)})

    # 5 dakikadan uzun süre running/pending kalan job'ları stale say → failed yap
    stale_cutoff = timezone.now() - timedelta(minutes=60)
    OAIPMHSearchJob.objects.filter(
        user=request.user,
        status__in=['pending', 'running'],
        created_at__lt=stale_cutoff,
    ).update(status='failed', error_message='Sunucu yeniden başlatıldı veya zaman aşımı.')

    active_job = OAIPMHSearchJob.objects.filter(
        user=request.user,
        status__in=['pending', 'running'],
    ).order_by('-created_at').first()

    return render(request, 'oaipmh/landing.html', {
        'keyword_form': OAIPMHKeywordForm(),
        'browse_form': OAIPMHBrowseForm(),
        'daily_limit': daily_limit,
        'remaining': remaining,
        'active_job_id': str(active_job.id) if active_job else None,
        'active_job_type': active_job.search_type if active_job else None,
        'seo_guide': TARAMA_SEO_CONTENT.get('oaipmh'),
    })


@login_required
@require_GET
def oaipmh_job_status(request, job_id):
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)

    data = {'status': job.status, 'total_results': job.total_results}

    if job.status == 'completed':
        data['demo_results'] = job.demo_results[:3]
        data['all_results_file_url'] = job.all_results_file_url
        data['demo_email_sent'] = job.demo_email_sent
        data['query_summary'] = job.get_query_summary()
    elif job.status == 'failed':
        data['error'] = job.error_message

    return JsonResponse(data)


@login_required
@require_POST
def oaipmh_send_demo_email(request, job_id):
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)

    if job.status != 'completed':
        return JsonResponse({'success': False, 'error': 'İş henüz tamamlanmadı.'}, status=400)
    if not job.demo_results:
        return JsonResponse({'success': False, 'error': 'Sonuç bulunamadı.'}, status=400)
    if job.demo_email_sent:
        return JsonResponse({'success': False, 'error': 'Demo email zaten gönderildi.'}, status=400)

    send_demo_email_async(job.id)
    return JsonResponse({'success': True})


@login_required
@require_POST
def oaipmh_cancel(request, job_id):
    """Devam eden taramayı iptal eder."""
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)
    if job.status in ('pending', 'running'):
        job.status = 'failed'
        job.error_message = 'Kullanıcı tarafından iptal edildi.'
        job.save(update_fields=['status', 'error_message'])
    return JsonResponse({'success': True})


def _download_filename(job, ext):
    """İndirme dosyası adı: üniversite_adı + zaman_damgası."""
    import re
    ts = (job.completed_at or job.created_at).strftime('%Y%m%d_%H%M%S')
    if job.search_type == 'browse' and job.university:
        uni = re.sub(r'[^\w\s-]', '', job.university.name).strip().replace(' ', '_')
    elif job.keyword:
        uni = re.sub(r'[^\w\s-]', '', job.keyword).strip().replace(' ', '_')[:30]
    else:
        uni = 'tez'
    return f"{uni}_{ts}.{ext}"


@login_required
@require_GET
def oaipmh_download(request, job_id):
    """S3'teki TXT dosyasını proxy ile indirtir (tarayıcıda açmaz)."""
    import requests as req_lib
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)

    if not job.all_results_file_url:
        raise Http404

    try:
        s3_resp = req_lib.get(job.all_results_file_url, timeout=30, stream=True)
        s3_resp.raise_for_status()
    except Exception:
        raise Http404

    response = HttpResponse(
        s3_resp.content,
        content_type='text/plain; charset=utf-8',
    )
    fname = _download_filename(job, 'txt')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
@require_GET
def oaipmh_download_excel(request, job_id):
    """demo_results JSON'ından Excel (.xlsx) üretir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)
    records = job.demo_results or []
    if not records:
        raise Http404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Üniversite Tez Sonuçları'

    headers = ['No', 'Başlık', 'Yazarlar', 'Yıl', 'Üniversite', 'Tür', 'Konu', 'Özet', 'Link']
    ws.append(headers)
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for i, r in enumerate(records, 1):
        ws.append([
            i,
            r.get('title', ''),
            r.get('authors', ''),
            r.get('year', ''),
            r.get('university', ''),
            r.get('type', ''),
            r.get('subject', ''),
            r.get('abstract', ''),
            r.get('link', ''),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = _download_filename(job, 'xlsx')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required
def oaipmh_order_page(request, job_id):
    _feature_check()
    job = get_object_or_404(OAIPMHSearchJob, id=job_id, user=request.user)

    if job.status != 'completed' or not job.demo_results:
        return redirect('oaipmh:landing')
    # Demo zaten tüm sonuçları kapsıyorsa sipariş anlamsız
    if job.total_results <= len(job.demo_results):
        return redirect('oaipmh:landing')

    existing_order = OAIPMHOrder.objects.filter(search_job=job).first()
    total_price = OAIPMHOrder.calculate_price(job.total_results)

    if request.method == 'POST':
        if existing_order:
            return redirect('oaipmh:order_page', job_id=job_id)
        form = OAIPMHOrderForm(request.POST)
        if form.is_valid():
            count = min(form.cleaned_data['abstract_count'], job.total_results)
            price = OAIPMHOrder.calculate_price(count)
            OAIPMHOrder.objects.create(
                user=request.user,
                search_job=job,
                abstract_count=count,
                total_price=price,
                payment_note=form.cleaned_data.get('payment_note', ''),
            )
            return redirect('oaipmh:order_page', job_id=job_id)
    else:
        form = OAIPMHOrderForm(initial={'abstract_count': min(job.total_results, 100)})

    return render(request, 'oaipmh/order.html', {
        'job': job,
        'form': form,
        'existing_order': existing_order,
        'total_price': total_price,
        'iban_info': IBAN_INFO,
    })
